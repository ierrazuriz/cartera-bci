#!/usr/bin/env python3
"""
Cartera BCI - EL LTDA (76.677.950-6) y EMF SPA (77.209.686-0)
Generador de Excel basado en cartola del 27/03/2026.
Ejecutar diariamente: actualizar precios en hoja PRECIOS -> todo se recalcula.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

CARPETA = os.path.dirname(os.path.abspath(__file__))
ARCHIVO = os.path.join(CARPETA, "Cartera BCI - EL y EMF SPA.xlsx")

# ── Colores ──────────────────────────────────────────────────────────────────
C_NAV      = "003366"   # azul oscuro encabezado principal
C_BLUE     = "2E75B6"   # azul encabezado sección
C_LBLUE    = "BDD7EE"   # azul claro fila par
C_VLIGHT   = "DEEAF1"   # azul muy claro alternado
C_WHITE    = "FFFFFF"
C_YELLOW   = "FFFFC0"   # celdas de entrada
C_GRAY     = "F2F2F2"
C_GREEN    = "C6EFCE"
C_RED      = "FFC7CE"
C_DGRAY    = "595959"

# ── Helpers ──────────────────────────────────────────────────────────────────
def fill(color):
    return PatternFill("solid", fgColor=color)

def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="thin")
    return Border(bottom=s)

def w(ws, row, col, val=None, formula=None, bold=False, sz=10, color="000000",
      bg=None, ha="left", wrap=False, fmt=None, italic=False):
    """Escribe una celda con formato."""
    c = ws.cell(row=row, column=col)
    c.value = formula if formula else val
    c.font  = font(bold=bold, size=sz, color=color, italic=italic)
    c.alignment = align(h=ha, wrap=wrap)
    if bg:  c.fill = fill(bg)
    if fmt: c.number_format = fmt
    return c

def hdr(ws, row, items, bg=C_NAV, tc=C_WHITE, sz=10, h=22):
    """Escribe una fila de encabezado."""
    ws.row_dimensions[row].height = h
    for col, val in items:
        c = ws.cell(row=row, column=col)
        c.value = val
        c.font  = font(bold=True, size=sz, color=tc)
        c.fill  = fill(bg)
        c.alignment = align(h="center", v="center", wrap=True)
        c.border = border()

def section(ws, row, title, ncols=9, bg=C_BLUE, start_col=2):
    """Fila de título de sección."""
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=start_col+ncols-1)
    c = ws.cell(row=row, column=start_col)
    c.value = title
    c.font  = font(bold=True, size=11, color=C_WHITE)
    c.fill  = fill(bg)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[row].height = 22

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# ════════════════════════════════════════════════════════════════════════════
# DATOS: posiciones al 27/03/2026
# ════════════════════════════════════════════════════════════════════════════

# EL LTDA - Acciones (nemotécnico, nombre, cant_activo, cant_pasivo, precio_cartola)
EL_ACCIONES = [
    ("ABC",      "ABC S.A.",                  23_210_430,         0,   12.09),
    ("AGUAS-A",  "Aguas Andinas S.A.",          1_939_069,  -120_000,  348.00),
    ("CENCOSUD", "Cencosud S.A.",                  86_229,         0, 2_340.00),
    ("CHILE",    "Banco de Chile",             6_000_000, -1_000_000,  168.00),
    ("COPEC",    "Empresas Copec S.A.",            21_055,         0, 6_290.00),
    ("ENELAM",   "Enel Americas S.A.",         13_158_102,         0,   77.25),
    ("LTM",      "Latam Airlines Group S.A.", 104_595_390, -5_000_000,  22.83),
]

# EL LTDA - CFI (nemotécnico, nombre, cantidad, precio_compra, precio_cartola)
EL_CFI = [
    ("CFIARRAA-E", "CFI Arraa Serie E",           3_801, 47_777.8556, 53_279.4282),
    ("CFIMRCLP",   "Moneda Renta CLP FI, Serie A",11_172, 19_592.0000, 20_090.7100),
    ("CFITRIPT-E", "CFI Tript Serie E",            1_471, 13_280.7610, 12_000.0000),
]

# EL LTDA - Simultáneas (instrumento, cantidad, f_venta, monto_venta, f_compra, monto_compra)
EL_SIM = [
    ("AGUAS-A",   558_600, date(2026,3,26), 193_834_200, date(2026,4,23), 194_684_501),
    ("CHILE",     238_830, date(2026,3,20),  42_370_830, date(2026,4,17),  42_556_688),
    ("COPEC",       2_286, date(2026,3,26),  14_401_800, date(2026,4,23),  14_464_976),
    ("ENELAM", 7_905_720,  date(2026,3, 6), 639_335_576, date(2026,4, 1), 641_939_721),
    ("ENELAM", 1_816_050,  date(2026,3,12), 139_091_270, date(2026,4, 9), 139_701_462),
    ("ENELAM",    75_000,  date(2026,3,16),   6_000_000, date(2026,4,15),   6_028_200),
    ("ENELAM", 1_061_332,  date(2026,3,20),  85_437_226, date(2026,4,17),  85_811_982),
    ("LTM",   15_310_300,  date(2026,3, 6), 335_601_776, date(2026,4, 1), 336_968_986),
    ("LTM",   69_285_090,  date(2026,3,12),1_525_657_682,date(2026,4, 9),1_532_350_621),
    ("LTM",   20_000_000,  date(2026,3,20), 446_200_000, date(2026,4,17), 447_878_000),
]

# EMF SPA - CFI
EMF_CFI = [
    ("CFIARRAA-E", "CFI Arraa Serie E", 500, 47_154.0000, 53_279.4282),
]

# EMF SPA - Forwards
# (folio, tipo C/V, usd, tc_fwd, f_inicio, f_termino)
EMF_FWD = [
    (1832537, "C", 500_000, 908.20, date(2026,3,23), date(2026,4, 2)),
    (1833221, "V", 500_000, 920.90, date(2026,3,26), date(2026,4, 8)),
    (1833037, "V", 500_000, 916.40, date(2026,3,25), date(2026,4, 9)),
    (1832341, "V", 500_000, 915.55, date(2026,3,20), date(2026,4,10)),
    (1832343, "V", 500_000, 921.35, date(2026,3,20), date(2026,4,10)),
    (1832345, "V", 500_000, 926.25, date(2026,3,20), date(2026,4,10)),
    (1831846, "V", 500_000, 912.98, date(2026,3,18), date(2026,4, 2)),
]

# Mapa nemotécnico -> fila en hoja PRECIOS (col E = precio hoy)
# Fila 14..22 en PRECIOS (ver construcción más abajo)
PRECIOS_ROW = {
    "ABC":        14,
    "AGUAS-A":    15,
    "CENCOSUD":   16,
    "CHILE":      17,
    "COPEC":      18,
    "ENELAM":     19,
    "LTM":        20,
    "CFIARRAA-E": 21,
    "CFIMRCLP":   22,
    "CFITRIPT-E": 23,
}
P_UF  = "C8"   # UF en PRECIOS
P_USD = "C9"   # USD en PRECIOS
P_EUR = "C10"  # EUR en PRECIOS
P_HOY = "C4"   # Fecha de valorización en PRECIOS

def precio_ref(nem):
    """Referencia a precio actual en hoja PRECIOS."""
    return f"PRECIOS!E{PRECIOS_ROW[nem]}"


# ════════════════════════════════════════════════════════════════════════════
# HOJA 1: PRECIOS
# ════════════════════════════════════════════════════════════════════════════
def build_precios(wb):
    ws = wb.active
    ws.title = "PRECIOS"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16

    # ── Título ────────────────────────────────────────────────────────────
    merge(ws,1,2,1,6)
    w(ws,1,2,"PANEL DE PRECIOS DIARIOS", bold=True, sz=14, color=C_WHITE,
      bg=C_NAV, ha="center")
    ws.row_dimensions[1].height = 34

    # ── Info fecha ────────────────────────────────────────────────────────
    w(ws,3,2,"Cartola base:", bold=True)
    w(ws,3,3,date(2026,3,27), fmt="DD/MM/YYYY", ha="center")
    w(ws,4,2,"Fecha valorización:", bold=True)
    c = ws.cell(4,3); c.value = "=TODAY()"; c.number_format = "DD/MM/YYYY"
    c.font = font(bold=True, color="00B050"); c.alignment = align(h="center")

    # ── Valores de referencia ─────────────────────────────────────────────
    section(ws, 6, "VALORES DE REFERENCIA  (actualizar diariamente)", ncols=4)

    hdr(ws, 7, [(2,"Indicador"),(3,"Valor Base\n(cartola 27/03)"),(4,"Valor Hoy"),(5,"")],
        bg=C_BLUE, h=30)

    ref_data = [
        (8,  "UF ($/UF)",  39_841.72),
        (9,  "USD/CLP",      923.25),
        (10, "EUR/CLP",    1_064.63),
    ]
    for r, label, base in ref_data:
        w(ws, r, 2, label, bold=True, bg=C_GRAY)
        w(ws, r, 3, base, fmt="#,##0.00", ha="right", bg=C_GRAY)
        c = ws.cell(r, 4); c.value = base
        c.number_format = "#,##0.00"; c.alignment = align(h="right")
        c.fill = fill(C_YELLOW); c.border = border()

    # ── Precios instrumentos ──────────────────────────────────────────────
    section(ws, 12, "PRECIOS ACCIONES / CFI  (actualizar con precio de cierre)", ncols=4)

    hdr(ws, 13, [(2,"Nemotécnico"),(3,"Nombre"),(4,"Precio Cartola\n27/03/2026"),
                 (5,"Precio Hoy ★"),(6,"Var. %")], bg=C_BLUE, h=35)

    instruments = [
        ("ABC",        "ABC S.A.",                   12.09,    "#,##0.0000"),
        ("AGUAS-A",    "Aguas Andinas S.A.",         348.00,   "#,##0.00"),
        ("CENCOSUD",   "Cencosud S.A.",            2_340.00,   "#,##0.00"),
        ("CHILE",      "Banco de Chile",             168.00,   "#,##0.00"),
        ("COPEC",      "Empresas Copec S.A.",      6_290.00,   "#,##0.00"),
        ("ENELAM",     "Enel Americas S.A.",          77.25,   "#,##0.00"),
        ("LTM",        "Latam Airlines G. S.A.",      22.83,   "#,##0.0000"),
        ("CFIARRAA-E", "CFI Arraa Serie E",      53_279.4282,  "#,##0.0000"),
        ("CFIMRCLP",   "Moneda Renta CLP, Serie A",20_090.71,  "#,##0.0000"),
        ("CFITRIPT-E", "CFI Tript Serie E",       12_000.00,   "#,##0.0000"),
    ]

    for i, (nem, nombre, precio_base, fmt) in enumerate(instruments):
        r = 14 + i
        bg = C_GRAY if i % 2 == 0 else None
        w(ws, r, 2, nem,    bold=True, bg=bg)
        w(ws, r, 3, nombre, bg=bg, sz=9)
        w(ws, r, 4, precio_base, fmt=fmt, ha="right", bg=bg)
        # Precio hoy - entrada de usuario
        c = ws.cell(r, 5); c.value = precio_base
        c.number_format = fmt; c.alignment = align(h="right")
        c.fill = fill(C_YELLOW); c.border = border()
        c.font = font(bold=True)
        # Variación %
        c2 = ws.cell(r, 6)
        c2.value = f"=IF(D{r}<>0,(E{r}-D{r})/D{r},0)"
        c2.number_format = "0.00%"; c2.alignment = align(h="right")
        if bg: c2.fill = fill(bg)

    # Nota
    w(ws, 25, 2, "★ Columna amarilla = precio de cierre a ingresar manualmente cada día.",
      sz=9, color=C_DGRAY, italic=True)
    w(ws, 26, 2, "★ UF/USD/EUR también se actualizan en la sección de Valores de Referencia.",
      sz=9, color=C_DGRAY, italic=True)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# HOJA 2: EL LTDA
# ════════════════════════════════════════════════════════════════════════════
def build_el(wb):
    ws = wb.create_sheet("EL LTDA (76.677.950-6)")
    ws.sheet_view.showGridLines = False

    cw = {"A":3,"B":24,"C":14,"D":14,"E":14,"F":14,"G":14,"H":16,"I":18,"J":18}
    for col, width in cw.items():
        ws.column_dimensions[col].width = width

    # ── Encabezado ────────────────────────────────────────────────────────
    merge(ws,1,2,1,10)
    w(ws,1,2,"INFORME DE CARTERA — ASESORIAS E INVERSIONES EL LTDA.",
      bold=True, sz=13, color=C_WHITE, bg=C_NAV, ha="center")
    ws.row_dimensions[1].height = 32

    info = [(2,"RUT:","76.677.950-6"),(3,"Cliente:","ASESORIAS E INVERSIONES EL LTDA."),
            (4,"Dirección:","Av. Candelaria Goyenechea 3900 Of. 501, Vitacura"),
            (5,"Cartola base:","27/03/2026 — Período 01/03/2026 al 27/03/2026")]
    for r, lbl, val in info:
        w(ws,r,2,lbl,bold=True)
        w(ws,r,3,val); merge(ws,r,3,r,6)

    w(ws,2,8,"UF:",bold=True,ha="right"); merge(ws,2,8,2,8)
    c=ws.cell(2,9); c.value=f"=PRECIOS!{P_UF}"; c.number_format="#,##0.00"; c.alignment=align(h="right")
    w(ws,3,8,"USD/CLP:",bold=True,ha="right")
    c=ws.cell(3,9); c.value=f"=PRECIOS!{P_USD}"; c.number_format="#,##0.00"; c.alignment=align(h="right")
    w(ws,4,8,"Fecha val.:",bold=True,ha="right")
    c=ws.cell(4,9); c.value=f"=PRECIOS!{P_HOY}"; c.number_format="DD/MM/YYYY"
    c.font=font(bold=True,color="00B050"); c.alignment=align(h="right")

    # ════════════════════════════════════
    # SECCIÓN: DETALLE ACCIONES (fila 8)
    # ════════════════════════════════════
    section(ws, 8, "DETALLE CARTERA — ACCIONES", ncols=9)

    hdr(ws, 9, [
        (2,"Instrumento"),(3,"Nombre"),(4,"Cant.\nActivo"),(5,"Cant.\nPasivo"),
        (6,"Precio\nCartola"),(7,"Precio\nHoy"),(8,"Valor\nActivo CLP"),
        (9,"Valor\nPasivo CLP"),(10,"VALOR\nNETO CLP")
    ], bg=C_BLUE, h=40)

    acc_r_first = 10
    for i,(nem,nombre,cant_a,cant_p,precio_c) in enumerate(EL_ACCIONES):
        r = acc_r_first + i
        bg = C_LBLUE if i%2==0 else None
        w(ws,r,2,nem,   bold=True, bg=bg)
        w(ws,r,3,nombre,bg=bg, sz=9)
        w(ws,r,4,cant_a, fmt="#,##0", ha="right", bg=bg)
        if cant_p != 0:
            w(ws,r,5,cant_p, fmt="#,##0", ha="right", bg=bg)
        else:
            w(ws,r,5,"—", ha="center", bg=bg)
        w(ws,r,6,precio_c, fmt="#,##0.0000", ha="right", bg=bg)
        # Precio actual
        c=ws.cell(r,7); c.value=f"={precio_ref(nem)}"
        c.number_format="#,##0.0000"; c.alignment=align(h="right")
        c.fill=fill(C_YELLOW); c.font=font(bold=True)
        # Valor activo
        c=ws.cell(r,8); c.value=f"=D{r}*G{r}"
        c.number_format="#,##0"; c.alignment=align(h="right")
        if bg: c.fill=fill(bg)
        # Valor pasivo
        c=ws.cell(r,9)
        c.value = f"=E{r}*G{r}" if cant_p != 0 else 0
        c.number_format="#,##0"; c.alignment=align(h="right")
        if bg: c.fill=fill(bg)
        # Neto
        c=ws.cell(r,10); c.value=f"=H{r}+I{r}"
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True)
        if bg: c.fill=fill(bg)

    acc_r_last = acc_r_first + len(EL_ACCIONES) - 1
    tot_acc_r  = acc_r_last + 1
    merge(ws,tot_acc_r,2,tot_acc_r,7)
    w(ws,tot_acc_r,2,"TOTAL ACCIONES", bold=True, bg=C_BLUE, color=C_WHITE, ha="center")
    for col,(formula) in [(8,f"=SUM(H{acc_r_first}:H{acc_r_last})"),
                           (9,f"=SUM(I{acc_r_first}:I{acc_r_last})"),
                           (10,f"=SUM(J{acc_r_first}:J{acc_r_last})")]:
        c=ws.cell(tot_acc_r,col); c.value=formula
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True,color=C_WHITE); c.fill=fill(C_BLUE)

    # Guardamos referencias para el resumen
    TOT_ACC_NET = f"J{tot_acc_r}"  # valor neto acciones
    TOT_ACC_ACT = f"H{tot_acc_r}"

    # ════════════════════════════════════
    # SECCIÓN: DETALLE CFI
    # ════════════════════════════════════
    cfi_sec_r = tot_acc_r + 2
    section(ws, cfi_sec_r, "DETALLE CARTERA — CFI", ncols=9)

    cfi_hdr_r = cfi_sec_r + 1
    hdr(ws, cfi_hdr_r, [
        (2,"Instrumento"),(3,"Nombre"),(4,"Cantidad"),
        (5,"Precio\nCompra"),(6,"Precio\nCartola"),(7,"Precio\nHoy"),
        (8,"Valor Mercado\nCLP"),(9,"Div.\nRecibidos"),(10,"")
    ], bg=C_BLUE, h=40)

    cfi_r_first = cfi_hdr_r + 1
    for i,(nem,nombre,cant,p_compra,p_cartola) in enumerate(EL_CFI):
        r = cfi_r_first + i
        bg = C_LBLUE if i%2==0 else None
        w(ws,r,2,nem,   bold=True, bg=bg)
        w(ws,r,3,nombre,bg=bg, sz=9)
        w(ws,r,4,cant,   fmt="#,##0", ha="right", bg=bg)
        w(ws,r,5,p_compra, fmt="#,##0.0000", ha="right", bg=bg)
        w(ws,r,6,p_cartola,fmt="#,##0.0000", ha="right", bg=bg)
        c=ws.cell(r,7); c.value=f"={precio_ref(nem)}"
        c.number_format="#,##0.0000"; c.alignment=align(h="right")
        c.fill=fill(C_YELLOW); c.font=font(bold=True)
        c=ws.cell(r,8); c.value=f"=D{r}*G{r}"
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True)
        if bg: c.fill=fill(bg)

    cfi_r_last = cfi_r_first + len(EL_CFI) - 1
    tot_cfi_r  = cfi_r_last + 1
    merge(ws,tot_cfi_r,2,tot_cfi_r,7)
    w(ws,tot_cfi_r,2,"TOTAL CFI", bold=True, bg=C_BLUE, color=C_WHITE, ha="center")
    c=ws.cell(tot_cfi_r,8); c.value=f"=SUM(H{cfi_r_first}:H{cfi_r_last})"
    c.number_format="#,##0"; c.alignment=align(h="right")
    c.font=font(bold=True,color=C_WHITE); c.fill=fill(C_BLUE)

    TOT_CFI = f"H{tot_cfi_r}"

    # ════════════════════════════════════
    # SECCIÓN: SIMULTÁNEAS
    # ════════════════════════════════════
    sim_sec_r = tot_cfi_r + 2
    section(ws, sim_sec_r, "SIMULTÁNEAS — COMPRAS A PLAZO (a valor de mercado)", ncols=9)

    sim_hdr_r = sim_sec_r + 1
    hdr(ws, sim_hdr_r, [
        (2,"Instrumento"),(3,"Cantidad"),(4,"Fecha\nVenta"),
        (5,"Monto\nVenta"),(6,"Fecha\nCompra Plazo"),(7,"Monto\nCompra Plazo"),
        (8,"Monto\nAmortizado Hoy"),(9,"Val. Mercado\nHoy"),(10,"RESULTADO")
    ], bg=C_BLUE, h=45)

    sim_r_first = sim_hdr_r + 1
    for i,(inst,cant,f_vta,m_vta,f_cpra,m_cpra) in enumerate(EL_SIM):
        r = sim_r_first + i
        bg = C_LBLUE if i%2==0 else None
        w(ws,r,2,inst, bold=True, bg=bg)
        w(ws,r,3,cant,  fmt="#,##0", ha="right", bg=bg)
        w(ws,r,4,f_vta, fmt="DD/MM/YYYY", ha="center", bg=bg)
        w(ws,r,5,m_vta, fmt="#,##0", ha="right", bg=bg)
        w(ws,r,6,f_cpra,fmt="DD/MM/YYYY", ha="center", bg=bg)
        w(ws,r,7,m_cpra,fmt="#,##0", ha="right", bg=bg)
        # Monto amortizado = lineal entre fecha venta y fecha compra
        c=ws.cell(r,8)
        c.value=(f"=E{r}+(G{r}-E{r})"
                 f"*MAX(0,MIN(PRECIOS!{P_HOY}-D{r},F{r}-D{r}))"
                 f"/(F{r}-D{r})")
        c.number_format="#,##0"; c.alignment=align(h="right")
        if bg: c.fill=fill(bg)
        # Valor de mercado
        c=ws.cell(r,9); c.value=f"=C{r}*{precio_ref(inst)}"
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.fill=fill(C_YELLOW)
        # Resultado
        c=ws.cell(r,10); c.value=f"=I{r}-H{r}"
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True)
        if bg: c.fill=fill(bg)

    sim_r_last = sim_r_first + len(EL_SIM) - 1
    tot_sim_r  = sim_r_last + 1
    merge(ws,tot_sim_r,2,tot_sim_r,7)
    w(ws,tot_sim_r,2,"TOTAL SIMULTÁNEAS", bold=True, bg=C_BLUE, color=C_WHITE, ha="center")
    for col, formula in [
        (8,f"=SUM(H{sim_r_first}:H{sim_r_last})"),
        (9,f"=SUM(I{sim_r_first}:I{sim_r_last})"),
        (10,f"=SUM(J{sim_r_first}:J{sim_r_last})")]:
        c=ws.cell(tot_sim_r,col); c.value=formula
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True,color=C_WHITE); c.fill=fill(C_BLUE)

    # Nota: simultáneas registradas como pasivo (compras a plazo = deuda)
    TOT_SIM_AMORT = f"H{tot_sim_r}"   # monto comprometido (negativo en patrimonio)
    TOT_SIM_VM    = f"I{tot_sim_r}"   # valor acciones comprometidas
    TOT_SIM_RES   = f"J{tot_sim_r}"   # resultado Mark-to-Market

    # ════════════════════════════════════
    # SECCIÓN: RESUMEN PATRIMONIAL
    # ════════════════════════════════════
    res_sec_r = tot_sim_r + 2
    section(ws, res_sec_r, "RESUMEN PATRIMONIAL — EL LTDA. (valorizado a mercado)", ncols=9)

    res_hdr_r = res_sec_r + 1
    hdr(ws, res_hdr_r, [(2,"Componente"),(9,"Monto CLP"),(10,"")], bg=C_BLUE)
    merge(ws,res_hdr_r,2,res_hdr_r,8)

    # Datos fijos de la cartola (saldo de caja no cambia sin nuevos movimientos)
    CAJA_BRUTA   = 555_976_435   # saldo final período
    OPS_LIQUIDAR = 205_558_872   # pagos por recibir/entregar (compras pendiente liquidación)
    CAJA_DISP    = 350_417_563   # = CAJA_BRUTA - OPS_LIQUIDAR

    componentes = [
        ("Caja Disponible CLP (según cartola 27/03)",     CAJA_DISP, True),
        ("  Saldo final período",                          CAJA_BRUTA, False),
        ("  Ops. por liquidar (compras pendientes)",      -OPS_LIQUIDAR, False),
        (None, None, None),
        ("Acciones — Valor Neto (activo - pasivo)",       None, True),
        ("CFI — Valor Mercado",                           None, True),
        ("Simultáneas — Posición comprometida (pasivo)", None, True),
        (None, None, None),
        ("PATRIMONIO TOTAL (CLP)",                        None, True),
        ("PATRIMONIO TOTAL (UF)",                         None, False),
        ("PATRIMONIO TOTAL (USD)",                        None, False),
    ]

    r = res_hdr_r + 1
    res_rows = {}  # nombre -> fila
    for label, val, bold_flag in componentes:
        if label is None:
            r += 1
            continue
        bg = None
        if bold_flag:
            bg = C_GRAY
        merge(ws,r,2,r,8)
        c=ws.cell(r,2); c.value=label
        c.font=font(bold=bold_flag, size=10 if bold_flag else 9,
                    color="000000" if bold_flag else C_DGRAY)
        c.alignment=align(); c.fill=fill(bg or C_WHITE)

        c2=ws.cell(r,9)
        if label.startswith("Caja Disponible"):
            c2.value = CAJA_DISP
            res_rows["caja"] = r
        elif label.startswith("  Saldo final"):
            c2.value = CAJA_BRUTA
        elif label.startswith("  Ops."):
            c2.value = -OPS_LIQUIDAR
        elif label.startswith("Acciones"):
            c2.value = f"={TOT_ACC_NET}"
            res_rows["acc"] = r
        elif label.startswith("CFI"):
            c2.value = f"={TOT_CFI}"
            res_rows["cfi"] = r
        elif label.startswith("Simultáneas"):
            # Simultáneas aparecen como pasivo: valor de mercado de acciones MENOS monto amortizado
            # El patrimonio ya incluye las acciones en garantía, y las simultáneas financian eso
            # Efecto neto = - monto amortizado (deuda de repos)
            c2.value = f"=-{TOT_SIM_AMORT}"
            res_rows["sim"] = r
        elif label.startswith("PATRIMONIO TOTAL (CLP)"):
            r_pat = r
            c2.value = (f"=I{res_rows.get('caja',0)}"
                        f"+I{res_rows.get('acc',0)}"
                        f"+I{res_rows.get('cfi',0)}"
                        f"+I{res_rows.get('sim',0)}")
            c2.font=font(bold=True,size=12,color=C_WHITE)
            c2.fill=fill(C_NAV)
            ws.cell(r,2).fill=fill(C_NAV)
            ws.cell(r,2).font=font(bold=True,size=12,color=C_WHITE)
            res_rows["pat_clp"] = r
        elif label.startswith("PATRIMONIO TOTAL (UF)"):
            c2.value = f"=I{res_rows.get('pat_clp',0)}/PRECIOS!{P_UF}"
            c2.number_format = "#,##0.00"
        elif label.startswith("PATRIMONIO TOTAL (USD)"):
            c2.value = f"=I{res_rows.get('pat_clp',0)}/PRECIOS!{P_USD}"
            c2.number_format = "#,##0.00"
        if isinstance(c2.value, (int,float)):
            c2.number_format = "#,##0"
        elif c2.value and str(c2.value).startswith("="):
            if "UF" not in label and "USD" not in label:
                c2.number_format = "#,##0"
        c2.alignment = align(h="right")
        if bold_flag and "PATRIMONIO" not in label:
            c2.font = font(bold=True)
        r += 1

    # Nota MtM simultáneas
    w(ws, r+1, 2,
      f"Nota: Resultado MtM simultáneas al día de hoy = {TOT_SIM_RES} "
      f"(diferencia entre valor de mercado acciones y monto amortizado del repo).",
      sz=9, italic=True, color=C_DGRAY)
    merge(ws, r+1, 2, r+1, 10)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# HOJA 3: EMF SPA
# ════════════════════════════════════════════════════════════════════════════
def build_emf(wb):
    ws = wb.create_sheet("EMF SPA (77.209.686-0)")
    ws.sheet_view.showGridLines = False

    cw = {"A":3,"B":26,"C":14,"D":14,"E":14,"F":14,"G":14,"H":16,"I":18,"J":14}
    for col, width in cw.items():
        ws.column_dimensions[col].width = width

    # ── Encabezado ────────────────────────────────────────────────────────
    merge(ws,1,2,1,10)
    w(ws,1,2,"INFORME DE CARTERA — EMF SPA",
      bold=True, sz=13, color=C_WHITE, bg=C_NAV, ha="center")
    ws.row_dimensions[1].height = 32

    info = [(2,"RUT:","77.209.686-0"),(3,"Cliente:","EMF SPA"),
            (4,"Dirección:","Av. Candelaria Goyenechea 3900 Of. 501, Vitacura"),
            (5,"Cartola base:","27/03/2026 — Período 01/03/2026 al 27/03/2026")]
    for r, lbl, val in info:
        w(ws,r,2,lbl,bold=True)
        w(ws,r,3,val); merge(ws,r,3,r,6)

    w(ws,2,8,"UF:",bold=True,ha="right")
    c=ws.cell(2,9); c.value=f"=PRECIOS!{P_UF}"; c.number_format="#,##0.00"; c.alignment=align(h="right")
    w(ws,3,8,"USD/CLP:",bold=True,ha="right")
    c=ws.cell(3,9); c.value=f"=PRECIOS!{P_USD}"; c.number_format="#,##0.00"; c.alignment=align(h="right")
    w(ws,4,8,"Fecha val.:",bold=True,ha="right")
    c=ws.cell(4,9); c.value=f"=PRECIOS!{P_HOY}"; c.number_format="DD/MM/YYYY"
    c.font=font(bold=True,color="00B050"); c.alignment=align(h="right")

    # ════════════════════════════════════
    # SECCIÓN: CFI
    # ════════════════════════════════════
    section(ws, 7, "DETALLE CARTERA — CFI", ncols=9)
    hdr(ws, 8, [
        (2,"Instrumento"),(3,"Nombre"),(4,"Cantidad"),
        (5,"Precio\nCompra"),(6,"Precio\nCartola"),(7,"Precio\nHoy"),
        (8,"Valor Mercado\nCLP"),(9,""),(10,"")
    ], bg=C_BLUE, h=40)

    r = 9
    for i,(nem,nombre,cant,p_comp,p_cart) in enumerate(EMF_CFI):
        bg = C_LBLUE if i%2==0 else None
        w(ws,r,2,nem,   bold=True,bg=bg)
        w(ws,r,3,nombre,bg=bg,sz=9)
        w(ws,r,4,cant,   fmt="#,##0",ha="right",bg=bg)
        w(ws,r,5,p_comp, fmt="#,##0.0000",ha="right",bg=bg)
        w(ws,r,6,p_cart, fmt="#,##0.0000",ha="right",bg=bg)
        c=ws.cell(r,7); c.value=f"={precio_ref(nem)}"
        c.number_format="#,##0.0000"; c.alignment=align(h="right")
        c.fill=fill(C_YELLOW); c.font=font(bold=True)
        c=ws.cell(r,8); c.value=f"=D{r}*G{r}"
        c.number_format="#,##0"; c.alignment=align(h="right"); c.font=font(bold=True)
        if bg: c.fill=fill(bg)
        r += 1

    cfi_r_last = r - 1
    cfi_r_first = 9
    tot_cfi_emf = r
    merge(ws,tot_cfi_emf,2,tot_cfi_emf,7)
    w(ws,tot_cfi_emf,2,"TOTAL CFI",bold=True,bg=C_BLUE,color=C_WHITE,ha="center")
    c=ws.cell(tot_cfi_emf,8); c.value=f"=SUM(H{cfi_r_first}:H{cfi_r_last})"
    c.number_format="#,##0"; c.alignment=align(h="right")
    c.font=font(bold=True,color=C_WHITE); c.fill=fill(C_BLUE)
    TOT_CFI_EMF = f"H{tot_cfi_emf}"

    # ════════════════════════════════════
    # SECCIÓN: FORWARDS
    # ════════════════════════════════════
    fwd_sec_r = tot_cfi_emf + 2
    section(ws, fwd_sec_r, "DETALLE FORWARDS USD/CLP — Seguros de Cambio", ncols=9)

    fwd_hdr_r = fwd_sec_r + 1
    hdr(ws, fwd_hdr_r, [
        (2,"Folio"),(3,"Tipo\nC/V"),(4,"USD\nNominal"),(5,"TC\nForward"),
        (6,"F. Inicio"),(7,"F. Término"),(8,"TC Spot\nHoy"),
        (9,"Resultado\nHoy (CLP)"),(10,"Vence\nen días")
    ], bg=C_BLUE, h=45)

    fwd_r_first = fwd_hdr_r + 1
    r = fwd_r_first
    for i,(folio,tipo,usd,tc_fwd,f_ini,f_term) in enumerate(EMF_FWD):
        bg = C_LBLUE if i%2==0 else None
        w(ws,r,2,folio,ha="center",bg=bg)
        w(ws,r,3,tipo, ha="center",bold=True,bg=bg,
          color="00B050" if tipo=="C" else "C00000")
        w(ws,r,4,usd,  fmt="#,##0",ha="right",bg=bg)
        w(ws,r,5,tc_fwd,fmt="#,##0.0000",ha="right",bg=bg)
        w(ws,r,6,f_ini, fmt="DD/MM/YYYY",ha="center",bg=bg)
        w(ws,r,7,f_term,fmt="DD/MM/YYYY",ha="center",bg=bg)
        # TC Spot hoy (referencia a PRECIOS)
        c=ws.cell(r,8); c.value=f"=PRECIOS!{P_USD}"
        c.number_format="#,##0.0000"; c.alignment=align(h="right")
        c.fill=fill(C_YELLOW)
        # Resultado: Compra = (Spot - FWD)*USD; Venta = (FWD - Spot)*USD
        if tipo == "C":
            formula = f"=(H{r}-E{r})*D{r}"
        else:
            formula = f"=(E{r}-H{r})*D{r}"
        c=ws.cell(r,9); c.value=formula
        c.number_format="#,##0"; c.alignment=align(h="right")
        c.font=font(bold=True)
        if bg: c.fill=fill(bg)
        # Días restantes
        c=ws.cell(r,10); c.value=f"=MAX(0,G{r}-PRECIOS!{P_HOY})"
        c.number_format="0"; c.alignment=align(h="right")
        if bg: c.fill=fill(bg)
        r += 1

    fwd_r_last = r - 1
    tot_fwd_r  = r
    merge(ws,tot_fwd_r,2,tot_fwd_r,8)
    w(ws,tot_fwd_r,2,"TOTAL FORWARDS (Resultado Hoy)", bold=True, bg=C_BLUE, color=C_WHITE, ha="center")
    c=ws.cell(tot_fwd_r,9); c.value=f"=SUM(I{fwd_r_first}:I{fwd_r_last})"
    c.number_format="#,##0"; c.alignment=align(h="right")
    c.font=font(bold=True,color=C_WHITE); c.fill=fill(C_BLUE)
    TOT_FWD = f"I{tot_fwd_r}"

    # Descalce
    compra_usd = sum(f[2] for f in EMF_FWD if f[1]=="C")
    venta_usd  = sum(f[2] for f in EMF_FWD if f[1]=="V")
    descalce   = compra_usd - venta_usd
    r2 = tot_fwd_r + 1
    w(ws,r2,2,f"Descalce neto: Compra {compra_usd:,.0f} USD / Venta {venta_usd:,.0f} USD / Neto {descalce:,.0f} USD",
      sz=9, italic=True, color=C_DGRAY)
    merge(ws,r2,2,r2,10)

    # ════════════════════════════════════
    # SECCIÓN: MOVIMIENTOS CAJA (HISTÓRICO FWD)
    # ════════════════════════════════════
    caj_sec_r = r2 + 2
    section(ws, caj_sec_r, "HISTORIAL MOVIMIENTOS CAJA — COMPENSACIONES FORWARD (marzo 2026)", ncols=9)

    caj_hdr_r = caj_sec_r + 1
    hdr(ws, caj_hdr_r, [(2,"Fecha"),(3,"Referencia"),(4,"Operación"),
                         (8,"Cargo CLP"),(9,"Abono CLP"),(10,"Saldo CLP")],
        bg=C_BLUE, h=30)
    merge(ws,caj_hdr_r,4,caj_hdr_r,7)

    CAJA_HIST = [
        (date(2026,3, 1),""       ,"Saldo inicial del período",   0,          7_761_120, 7_761_120),
        (date(2026,3,12),"1830442","Cargo comp. FWD",         9_045_000,          0,    -1_283_880),
        (date(2026,3,12),"1830441","Cargo comp. FWD",         8_970_000,          0,   -10_253_880),
        (date(2026,3,12),"1830440","Abono vcto. FWD comp.",          0,    9_995_000,     -258_880),
        (date(2026,3,12),"1830439","Abono vcto. FWD comp.",          0,   11_495_000,   11_236_120),
        (date(2026,3,13),"1830514","Abono vcto. FWD comp.",          0,    6_920_000,   18_156_120),
        (date(2026,3,13),"1829896","Cargo comp. FWD",         1_265_000,          0,   16_891_120),
        (date(2026,3,19),"1831499","Abono vcto. FWD comp.",          0,    2_305_000,   19_196_120),
        (date(2026,3,19),"1831351","Cargo comp. FWD",           865_000,          0,   18_331_120),
        (date(2026,3,19),"1831076","Cargo comp. FWD",         1_505_000,          0,   16_826_120),
        (date(2026,3,31),""       ,"Saldo final del período",         0,          0,   16_826_120),
    ]

    for i,(fecha,ref,oper,cargo,abono,saldo) in enumerate(CAJA_HIST):
        r3 = caj_hdr_r + 1 + i
        bg = C_GRAY if fecha==date(2026,3,1) or fecha==date(2026,3,31) else (C_LBLUE if i%2==0 else None)
        bold_r = fecha in (date(2026,3,1), date(2026,3,31))
        w(ws,r3,2,fecha,fmt="DD/MM/YYYY",ha="center",bold=bold_r,bg=bg)
        w(ws,r3,3,ref,  ha="center",bg=bg)
        merge(ws,r3,4,r3,7)
        w(ws,r3,4,oper, bold=bold_r,bg=bg)
        w(ws,r3,8, cargo if cargo>0 else "",fmt="#,##0",ha="right",bg=bg)
        w(ws,r3,9, abono if abono>0 else "",fmt="#,##0",ha="right",bg=bg)
        w(ws,r3,10,saldo,fmt="#,##0",ha="right",bold=bold_r,bg=bg)

    caja_saldo_emf = 16_826_120
    tot_caj_r = caj_hdr_r + 1 + len(CAJA_HIST)

    # ════════════════════════════════════
    # SECCIÓN: RESUMEN PATRIMONIAL EMF SPA
    # ════════════════════════════════════
    res_sec_r2 = tot_caj_r + 1
    section(ws, res_sec_r2, "RESUMEN PATRIMONIAL — EMF SPA (valorizado a mercado)", ncols=9)

    res_hdr_r2 = res_sec_r2 + 1
    hdr(ws, res_hdr_r2, [(2,"Componente"),(9,"Monto CLP"),(10,"")], bg=C_BLUE)
    merge(ws,res_hdr_r2,2,res_hdr_r2,8)

    pat_r = res_hdr_r2 + 1
    componentes_emf = [
        ("Caja Disponible CLP",              caja_saldo_emf, True),
        ("CFI — Valor Mercado",              None,           True),
        ("Forwards — Resultado MtM",         None,           True),
        ("PATRIMONIO TOTAL (CLP)",           None,           True),
        ("PATRIMONIO TOTAL (UF)",            None,           False),
        ("PATRIMONIO TOTAL (USD)",           None,           False),
    ]

    rows_emf = {}
    for label, val, bold_flag in componentes_emf:
        bg = C_GRAY if bold_flag else None
        merge(ws,pat_r,2,pat_r,8)
        c=ws.cell(pat_r,2); c.value=label
        c.font=font(bold=bold_flag,size=10 if bold_flag else 9)
        c.alignment=align(); c.fill=fill(bg or C_WHITE)
        c2=ws.cell(pat_r,9)
        if label.startswith("Caja"):
            c2.value=caja_saldo_emf; rows_emf["caja"]=pat_r
        elif label.startswith("CFI"):
            c2.value=f"={TOT_CFI_EMF}"; rows_emf["cfi"]=pat_r
        elif label.startswith("Forwards"):
            c2.value=f"={TOT_FWD}"; rows_emf["fwd"]=pat_r
        elif label.startswith("PATRIMONIO TOTAL (CLP)"):
            c2.value=(f"=I{rows_emf.get('caja',0)}"
                      f"+I{rows_emf.get('cfi',0)}"
                      f"+I{rows_emf.get('fwd',0)}")
            c2.font=font(bold=True,size=12,color=C_WHITE)
            c2.fill=fill(C_NAV)
            ws.cell(pat_r,2).fill=fill(C_NAV)
            ws.cell(pat_r,2).font=font(bold=True,size=12,color=C_WHITE)
            rows_emf["pat_clp"]=pat_r
        elif label.startswith("PATRIMONIO TOTAL (UF)"):
            c2.value=f"=I{rows_emf.get('pat_clp',0)}/PRECIOS!{P_UF}"
            c2.number_format="#,##0.00"
        elif label.startswith("PATRIMONIO TOTAL (USD)"):
            c2.value=f"=I{rows_emf.get('pat_clp',0)}/PRECIOS!{P_USD}"
            c2.number_format="#,##0.00"
        if isinstance(c2.value,(int,float)):
            c2.number_format="#,##0"
        elif c2.value and str(c2.value).startswith("=") and "UF" not in label and "USD" not in label:
            c2.number_format="#,##0"
        c2.alignment=align(h="right")
        if bold_flag and "PATRIMONIO" not in label:
            c2.font=font(bold=True)
        pat_r += 1

    # Nota forwards
    w(ws,pat_r+1,2,
      "Nota: Resultado forwards = mark-to-market al TC spot del día. "
      "Compra: ganancia cuando USD/CLP sube. Venta: ganancia cuando USD/CLP baja.",
      sz=9,italic=True,color=C_DGRAY)
    merge(ws,pat_r+1,2,pat_r+1,10)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    build_precios(wb)
    build_el(wb)
    build_emf(wb)

    # Congelar paneles
    for ws in wb.worksheets:
        ws.freeze_panes = "B2"

    wb.save(ARCHIVO)
    print(f"\nArchivo generado:\n  {ARCHIVO}\n")
    print("INSTRUCCIONES:")
    print("1. Abrir el archivo Excel.")
    print("2. Ir a hoja 'PRECIOS'.")
    print("3. Actualizar columna E (celdas amarillas) con precios de cierre del día.")
    print("4. Actualizar UF y USD/CLP si cambiaron.")
    print("5. Todo el resto se recalcula automáticamente.")
    print("6. Las hojas 'EL LTDA' y 'EMF SPA' muestran la cartera valorizada.")
    print("\nNOTA IMPORTANTE:")
    print("- La caja de EL LTDA está fija en CLP 350.417.563 (base cartola 27/03/2026).")
    print("  Actualizar manualmente en la hoja EL LTDA si hay nuevos movimientos.")
    print("- Las simultáneas que venzan deben eliminarse de la lista en el script.")
    print("- Los forwards de EMF SPA que venzan deben eliminarse de EMF_FWD.")


if __name__ == "__main__":
    main()
